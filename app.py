import streamlit as st
import pandas as pd
import openpyxl
import io
import zipfile
import os

st.set_page_config(page_title="WFS 货件工具", layout="wide", page_icon="📦")

st.title("📦 WFS 货件申请生成工具")

TEMPLATE_FILENAME = "WFS_M.xlsx"
template_exists = os.path.exists(TEMPLATE_FILENAME)

if not template_exists:
    st.error(f"系统缺失底层模板 {TEMPLATE_FILENAME}，请联系管理员维护。")
    st.stop()

# 侧边栏
with st.sidebar:
    inventory_file = st.file_uploader("1. 上传库存表 (Inventory)", type=['xlsx', 'csv'])
    packing_files = st.file_uploader("2. 上传装箱单 (支持多选/支持新老格式混排)", type=['xlsx', 'csv'], accept_multiple_files=True)
    
    unique_packing_files = []
    
    if packing_files:
        unique_packing_files = list({pack_file.name: pack_file for pack_file in packing_files}.values())
        
        if len(unique_packing_files) < len(packing_files):
            st.warning(f"⚠️ 已自动拦截 {len(packing_files) - len(unique_packing_files)} 份重复文件", icon="🪞")
        
        st.markdown("**✅ 实际有效装箱单列表：**")
        for f in unique_packing_files:
            st.caption(f"📄 {f.name}")

# 主界面逻辑
if inventory_file and unique_packing_files:
    if st.button("生成 WFS 申请表", type="primary"):
        with st.spinner('处理中...'):
            try:
                # 1. 读取库存表
                if inventory_file.name.endswith('.csv'):
                    inventory_df = pd.read_csv(inventory_file, dtype={'GTIN': str, 'SKU': str})
                else:
                    inventory_df = pd.read_excel(inventory_file, dtype={'GTIN': str, 'SKU': str})
                
                inventory_df['SKU'] = inventory_df['SKU'].astype(str).str.strip()
                
                summary_data = []
                generated_files = {}

                # 2. 遍历处理去重后的装箱单 (兼容新老格式)
                for pack_file in unique_packing_files:
                    is_new_format = False
                    xls = None
                    
                    # 格式嗅探：检查是否为包含特定分表的新格式Excel
                    if pack_file.name.endswith(('.xls', '.xlsx')):
                        try:
                            xls = pd.ExcelFile(pack_file)
                            if "装箱单统计" in xls.sheet_names and "调拨单-货物明细" in xls.sheet_names:
                                is_new_format = True
                        except Exception:
                            pass
                    
                    if is_new_format:
                        # ============== 新版装箱单逻辑 ==============
                        df_stats = pd.read_excel(xls, sheet_name="装箱单统计")
                        df_details = pd.read_excel(xls, sheet_name="调拨单-货物明细")
                        
                        # 单号：优先取调拨履约号(FA开头匹配WFS要求)，否则取调拨单号
                        if '调拨履约号' in df_details.columns and not df_details['调拨履约号'].dropna().empty:
                            order_id = df_details['调拨履约号'].dropna().iloc[0]
                        else:
                            order_id = df_details['调拨单号'].dropna().iloc[0] if '调拨单号' in df_details.columns and not df_details['调拨单号'].dropna().empty else "未知单号"
                        
                        # 箱数计算
                        box_count = len(df_stats['发货仓箱号'].dropna().unique())
                        
                        # 最重/最轻 重量计算
                        weights = pd.to_numeric(df_stats['参考计费重'], errors='coerce').dropna()
                        max_weight, min_weight = (weights.max(), weights.min()) if not weights.empty else (0, 0)
                        
                        # 数量透视汇总
                        df_details_clean = df_details.dropna(subset=['平台SKU'])
                        grouped_pack = df_details_clean.groupby('平台SKU', as_index=False)['期望出库数'].sum()
                        grouped_pack = grouped_pack.rename(columns={'期望出库数': '数量'})
                        total_qty = grouped_pack['数量'].sum()
                        
                    else:
                        # ============== 旧版装箱单逻辑 ==============
                        pack_file.seek(0) # 重置文件读取指针
                        if pack_file.name.endswith('.csv'):
                            pack_df = pd.read_csv(pack_file, header=3)
                        else:
                            pack_df = pd.read_excel(pack_file, header=3)

                        order_id = pack_df['出库单号'].dropna().iloc[0] if '出库单号' in pack_df.columns and not pack_df['出库单号'].dropna().empty else "未知单号"
                        box_count = len(pack_df['箱号'].dropna().unique())
                        
                        weights = pd.to_numeric(pack_df['重量'], errors='coerce').dropna()
                        max_weight, min_weight = (weights.max(), weights.min()) if not weights.empty else (0, 0)
                        
                        pack_df_clean = pack_df.dropna(subset=['平台SKU'])
                        grouped_pack = pack_df_clean.groupby('平台SKU', as_index=False)['数量'].sum()
                        total_qty = grouped_pack['数量'].sum()
                    
                    # 汇总数据收集（供页面报告展示）
                    summary_data.append({
                        "装箱单文件名": pack_file.name,
                        "出库单号": order_id,
                        "总数量": total_qty,
                        "装箱数": box_count,
                        "最重(kg)": f"{max_weight:.2f}",
                        "最轻(kg)": f"{min_weight:.2f}",
                        "文件格式": "新版(双分表)" if is_new_format else "旧版"
                    })
                    
                    # 3. 载入并填入官方模板 (共享写入逻辑)
                    wb = openpyxl.load_workbook(TEMPLATE_FILENAME)
                    ws = wb.active
                    
                    if ws.max_row >= 3:
                        for r in range(3, ws.max_row + 1):
                            for c in range(1, 10):
                                ws.cell(row=r, column=c).value = None
                    
                    start_row = 3
                    for idx, row in grouped_pack.iterrows():
                        sku = str(row['平台SKU']).strip()
                        qty = int(row['数量'])
                        
                        match = inventory_df[inventory_df['SKU'] == sku]
                        
                        if not match.empty:
                            gtin_val = match.iloc[0]['GTIN']
                            gtin = "" if pd.isna(gtin_val) else str(gtin_val).split('.')[0].strip().zfill(14)
                            item_name = str(match.iloc[0]['Item name']).strip()
                        else:
                            gtin, item_name = "未匹配到GTIN", "未匹配到产品名称"
                        
                        current_row = start_row + idx
                        
                        ws.cell(row=current_row, column=1, value="GTIN")
                        ws.cell(row=current_row, column=2, value=gtin)
                        ws.cell(row=current_row, column=3, value=sku)
                        ws.cell(row=current_row, column=4, value=item_name)
                        ws.cell(row=current_row, column=5, value=qty)
                        ws.cell(row=current_row, column=6, value=1)
                        ws.cell(row=current_row, column=7, value=qty)
                        ws.cell(row=current_row, column=8, value="No Prep Required")
                        ws.cell(row=current_row, column=9, value="No Prep Required")

                    output = io.BytesIO()
                    wb.save(output)
                    output.seek(0)
                    
                    generated_files[f"WFS_{order_id}.xlsx"] = output.read()
                    wb.close()

                # 4. 生成汇总并打包
                summary_df = pd.DataFrame(summary_data)
                summary_output = io.BytesIO()
                with pd.ExcelWriter(summary_output, engine='openpyxl') as writer:
                    summary_df.to_excel(writer, index=False, sheet_name='处理汇总')
                summary_output.seek(0)
                generated_files["处理汇总报告.xlsx"] = summary_output.read()

                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zip_file:
                    for file_name, data in generated_files.items():
                        zip_file.writestr(file_name, data)
                
                # 极简结果展示
                st.dataframe(summary_df, use_container_width=True)
                
                st.download_button(
                    label="下载压缩包 (ZIP)",
                    data=zip_buffer.getvalue(),
                    file_name="WFS_Export.zip",
                    mime="application/zip",
                    type="primary"
                )

            except Exception as e:
                st.error(f"数据格式错误，请检查传入的表格字段。报错详情：{str(e)}")
else:
    if not unique_packing_files:
        st.info("请在左侧上传必要的文件。")